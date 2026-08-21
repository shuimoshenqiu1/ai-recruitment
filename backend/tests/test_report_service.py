import sys
from unittest.mock import MagicMock

# Mock heavy dependencies that report_service doesn't need but __init__.py imports
for mod in ("fitz", "docx", "openai", "celery", "redis"):
    if mod not in sys.modules:
        sys.modules[mod] = MagicMock()

import io
import uuid
from dataclasses import dataclass
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.services.report_service import ReportService


@dataclass
class MockMatchResult:
    """模拟 MatchResult ORM 对象"""
    resume_id: uuid.UUID
    overall_score: Decimal
    skill_score: Decimal | None
    experience_score: Decimal | None
    education_score: Decimal | None
    soft_skill_score: Decimal | None
    grade: str | None
    recommendation: str | None
    model_used: str | None


@pytest.fixture
def report_service():
    return ReportService()


@pytest.fixture
def sample_matches():
    """生成测试用匹配结果列表"""
    return [
        MockMatchResult(
            resume_id=uuid.uuid4(),
            overall_score=Decimal("92.50"),
            skill_score=Decimal("95.00"),
            experience_score=Decimal("88.00"),
            education_score=Decimal("90.00"),
            soft_skill_score=Decimal("85.00"),
            grade="excellent",
            recommendation="强烈推荐面试",
            model_used="gpt-4o",
        ),
        MockMatchResult(
            resume_id=uuid.uuid4(),
            overall_score=Decimal("72.30"),
            skill_score=Decimal("68.00"),
            experience_score=Decimal("75.00"),
            education_score=Decimal("80.00"),
            soft_skill_score=Decimal("65.00"),
            grade="qualified",
            recommendation="可以安排面试",
            model_used="gpt-4o",
        ),
        MockMatchResult(
            resume_id=uuid.uuid4(),
            overall_score=Decimal("45.00"),
            skill_score=Decimal("35.00"),
            experience_score=Decimal("50.00"),
            education_score=Decimal("55.00"),
            soft_skill_score=Decimal("40.00"),
            grade="unqualified",
            recommendation="不建议面试",
            model_used="gpt-4o",
        ),
    ]


class TestGenerateMatchExcel:
    """测试增强版Excel导出"""

    def test_returns_bytesio(self, report_service, sample_matches):
        """应返回BytesIO对象"""
        result = report_service.generate_match_excel(sample_matches, "高级Python开发")
        assert isinstance(result, io.BytesIO)

    def test_valid_excel_file(self, report_service, sample_matches):
        """应生成合法的xlsx文件"""
        result = report_service.generate_match_excel(sample_matches, "高级Python开发")
        wb = load_workbook(result)
        assert wb.active is not None
        wb.close()

    def test_sheet_title_is_job_name(self, report_service, sample_matches):
        """Sheet名应为岗位名称"""
        result = report_service.generate_match_excel(sample_matches, "高级Python开发")
        wb = load_workbook(result)
        assert wb.active.title == "高级Python开发"
        wb.close()

    def test_sheet_title_truncated_at_31_chars(self, report_service, sample_matches):
        """Sheet名超过31字符应截断"""
        long_title = "这是一个非常非常非常长的岗位名称需要被截断处理以符合Excel要求"
        result = report_service.generate_match_excel(sample_matches, long_title)
        wb = load_workbook(result)
        assert len(wb.active.title) <= 31
        wb.close()

    def test_header_row_exists(self, report_service, sample_matches):
        """第一行应为表头"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "综合得分" in headers
        assert "等级" in headers
        assert "序号" in headers
        wb.close()

    def test_header_has_bold_font(self, report_service, sample_matches):
        """表头应为加粗字体"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        for cell in ws[1]:
            assert cell.font.bold is True
        wb.close()

    def test_header_has_background_fill(self, report_service, sample_matches):
        """表头应有背景色"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        first_header_cell = ws.cell(row=1, column=1)
        assert first_header_cell.fill.start_color.rgb is not None
        assert first_header_cell.fill.fill_type == "solid"
        wb.close()

    def test_data_rows_count(self, report_service, sample_matches):
        """数据行数应等于matches数量"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        # 表头 + 3数据行 + 1汇总行 = 5行
        assert ws.max_row == 5
        wb.close()

    def test_excellent_grade_has_green_fill(self, report_service, sample_matches):
        """excellent等级应有绿色背景"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        # 第2行（第一条数据）第8列是等级列
        grade_cell = ws.cell(row=2, column=8)
        assert grade_cell.fill.start_color.rgb == "00C6EFCE"
        wb.close()

    def test_qualified_grade_has_yellow_fill(self, report_service, sample_matches):
        """qualified等级应有黄色背景"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        grade_cell = ws.cell(row=3, column=8)
        assert grade_cell.fill.start_color.rgb == "00FFEB9C"
        wb.close()

    def test_unqualified_grade_has_red_fill(self, report_service, sample_matches):
        """unqualified等级应有红色背景"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        grade_cell = ws.cell(row=4, column=8)
        assert grade_cell.fill.start_color.rgb == "00FFC7CE"
        wb.close()

    def test_summary_row_has_average_score(self, report_service, sample_matches):
        """汇总行应有平均分"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        summary_row = len(sample_matches) + 2
        avg_cell = ws.cell(row=summary_row, column=3)
        expected_avg = round((92.50 + 72.30 + 45.00) / 3, 2)
        assert avg_cell.value == pytest.approx(expected_avg, abs=0.01)
        wb.close()

    def test_summary_row_has_grade_counts(self, report_service, sample_matches):
        """汇总行应有各等级人数统计"""
        result = report_service.generate_match_excel(sample_matches, "测试岗位")
        wb = load_workbook(result)
        ws = wb.active
        summary_row = len(sample_matches) + 2
        summary_cell = ws.cell(row=summary_row, column=8)
        assert "优秀: 1人" in summary_cell.value
        assert "合格: 1人" in summary_cell.value
        assert "不合格: 1人" in summary_cell.value
        assert "总计: 3人" in summary_cell.value
        wb.close()

    def test_empty_list_still_works(self, report_service):
        """空列表应能正常生成（无汇总行）"""
        result = report_service.generate_match_excel([], "空岗位")
        wb = load_workbook(result)
        ws = wb.active
        assert ws.max_row == 1  # 只有表头
        wb.close()

    def test_none_scores_handled(self, report_service):
        """部分分数为None时应正常处理"""
        matches = [
            MockMatchResult(
                resume_id=uuid.uuid4(),
                overall_score=Decimal("60.00"),
                skill_score=None,
                experience_score=None,
                education_score=Decimal("70.00"),
                soft_skill_score=None,
                grade="qualified",
                recommendation=None,
                model_used=None,
            )
        ]
        result = report_service.generate_match_excel(matches, "测试")
        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=4).value is None  # skill_score
        wb.close()


class TestGenerateSingleReportHtml:
    """测试单份匹配报告HTML生成"""

    @pytest.fixture
    def match_data(self):
        return {
            "overall_score": Decimal("85.50"),
            "skill_score": Decimal("90.00"),
            "experience_score": Decimal("80.00"),
            "education_score": Decimal("85.00"),
            "soft_skill_score": Decimal("75.00"),
            "grade": "excellent",
            "recommendation": "强烈推荐面试该候选人",
            "details": {
                "strengths": ["Python高级", "系统设计经验丰富", "团队协作好"],
                "weaknesses": ["缺少云原生经验"],
                "dimension_analysis": {
                    "技能匹配": "Python/FastAPI全栈开发经验5年+",
                    "经验匹配": "有大型项目主导经验",
                },
            },
        }

    @pytest.fixture
    def resume_data(self):
        return {
            "name": "张三",
            "email": "zhangsan@example.com",
            "phone": "13800138000",
            "education": [
                {"school": "清华大学", "major": "计算机科学", "degree": "硕士"},
            ],
            "experience": [
                {"company": "字节跳动", "position": "高级后端工程师", "duration": "2020-2024"},
            ],
        }

    @pytest.fixture
    def job_data(self):
        return {
            "title": "高级Python开发工程师",
            "department": "技术部",
            "level": "P7",
            "requirements": {"hard": ["Python", "FastAPI"], "soft": ["沟通能力"]},
        }

    def test_returns_html_string(self, report_service, match_data, resume_data, job_data):
        """应返回HTML字符串"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert isinstance(result, str)
        assert result.startswith("<!DOCTYPE html>")

    def test_contains_job_title(self, report_service, match_data, resume_data, job_data):
        """应包含岗位名称"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "高级Python开发工程师" in result

    def test_contains_candidate_name(self, report_service, match_data, resume_data, job_data):
        """应包含候选人姓名"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "张三" in result

    def test_contains_overall_score(self, report_service, match_data, resume_data, job_data):
        """应包含综合得分"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "85.5" in result

    def test_contains_grade_label(self, report_service, match_data, resume_data, job_data):
        """应包含等级标签"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "优秀" in result

    def test_contains_dimension_scores(self, report_service, match_data, resume_data, job_data):
        """应包含4个维度分数"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "90.0" in result  # skill
        assert "80.0" in result  # experience
        assert "85.0" in result  # education
        assert "75.0" in result  # soft skill

    def test_contains_strengths(self, report_service, match_data, resume_data, job_data):
        """应包含优势列表"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "Python高级" in result
        assert "系统设计经验丰富" in result

    def test_contains_weaknesses(self, report_service, match_data, resume_data, job_data):
        """应包含劣势列表"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "缺少云原生经验" in result

    def test_contains_recommendation(self, report_service, match_data, resume_data, job_data):
        """应包含推荐意见"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "强烈推荐面试该候选人" in result

    def test_contains_radar_data(self, report_service, match_data, resume_data, job_data):
        """应包含雷达图数据"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "radar-data" in result
        assert "data-labels" in result
        assert "data-values" in result

    def test_xss_prevention(self, report_service, match_data, job_data):
        """应防止XSS攻击"""
        malicious_resume = {
            "name": '<script>alert("xss")</script>',
            "email": "test@test.com",
            "phone": "123",
            "education": [],
            "experience": [],
        }
        result = report_service.generate_single_report_html(match_data, malicious_resume, job_data)
        assert "<script>" not in result
        assert "&lt;script&gt;" in result

    def test_handles_empty_details(self, report_service, resume_data, job_data):
        """details为空时应正常生成"""
        match_data = {
            "overall_score": Decimal("60.00"),
            "skill_score": None,
            "experience_score": None,
            "education_score": None,
            "soft_skill_score": None,
            "grade": "qualified",
            "recommendation": "",
            "details": None,
        }
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "<!DOCTYPE html>" in result
        assert "暂无推荐意见" in result

    def test_handles_empty_resume_data(self, report_service, match_data, job_data):
        """空简历数据应正常生成"""
        result = report_service.generate_single_report_html(match_data, {}, job_data)
        assert "未知候选人" in result

    def test_contains_education_info(self, report_service, match_data, resume_data, job_data):
        """应包含教育经历"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "清华大学" in result
        assert "计算机科学" in result

    def test_contains_work_experience(self, report_service, match_data, resume_data, job_data):
        """应包含工作经历"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "字节跳动" in result
        assert "高级后端工程师" in result

    def test_print_friendly(self, report_service, match_data, resume_data, job_data):
        """应包含打印友好的CSS"""
        result = report_service.generate_single_report_html(match_data, resume_data, job_data)
        assert "@media print" in result
