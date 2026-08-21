"""报告生成服务 - Excel导出增强 & 单份匹配报告"""

import io
from datetime import datetime, timezone
from decimal import Decimal
from html import escape
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- 颜色常量 ---
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXCELLENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
QUALIFIED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
UNQUALIFIED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SCORE_GREEN_FONT = Font(color="006100", bold=True)
SCORE_YELLOW_FONT = Font(color="9C5700", bold=True)
SCORE_RED_FONT = Font(color="9C0006", bold=True)
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUMMARY_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# --- 等级翻译 ---
GRADE_LABELS = {
    "excellent": "优秀",
    "qualified": "合格",
    "unqualified": "不合格",
}


def _get_grade_fill(grade: str | None) -> PatternFill | None:
    """根据等级返回对应的背景填充色。"""
    if grade == "excellent":
        return EXCELLENT_FILL
    elif grade == "qualified":
        return QUALIFIED_FILL
    elif grade == "unqualified":
        return UNQUALIFIED_FILL
    return None


def _get_score_font(score: float | Decimal | None) -> Font | None:
    """根据分数返回对应的字体颜色。"""
    if score is None:
        return None
    val = float(score)
    if val >= 80:
        return SCORE_GREEN_FONT
    elif val >= 60:
        return SCORE_YELLOW_FONT
    else:
        return SCORE_RED_FONT


def _auto_adjust_column_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """自动调整列宽（基于内容长度估算）。"""
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                # 中文字符约占2个宽度单位
                cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_length = max(max_length, cell_len)
        adjusted = max(min_width, min(max_length + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


class ReportService:
    """报告生成服务"""

    def generate_match_excel(
        self, matches: list[Any], job_title: str
    ) -> io.BytesIO:
        """生成带颜色标识的匹配结果Excel。

        Args:
            matches: MatchResult 对象列表（需有 resume_id, overall_score,
                     skill_score, experience_score, education_score,
                     soft_skill_score, grade, recommendation, model_used）
            job_title: 岗位名称，用作Sheet名

        Returns:
            包含Excel内容的BytesIO流
        """
        wb = Workbook()
        ws = wb.active
        # Sheet名最多31字符
        ws.title = job_title[:31] if job_title else "匹配结果"

        # --- 表头 ---
        headers = [
            "序号",
            "简历ID",
            "综合得分",
            "技能得分",
            "经验得分",
            "学历得分",
            "软技能得分",
            "等级",
            "推荐意见",
            "模型",
        ]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # --- 数据行 ---
        score_columns = [3, 4, 5, 6, 7]  # 分数列（1-indexed）
        grade_column = 8

        for idx, m in enumerate(matches, 1):
            row_data = [
                idx,
                str(m.resume_id),
                float(m.overall_score),
                float(m.skill_score) if m.skill_score else None,
                float(m.experience_score) if m.experience_score else None,
                float(m.education_score) if m.education_score else None,
                float(m.soft_skill_score) if m.soft_skill_score else None,
                GRADE_LABELS.get(m.grade, m.grade or ""),
                m.recommendation or "",
                m.model_used or "",
            ]
            ws.append(row_data)
            current_row = idx + 1  # +1 因为表头占第一行

            # 分数列颜色
            for col in score_columns:
                cell = ws.cell(row=current_row, column=col)
                font = _get_score_font(cell.value)
                if font:
                    cell.font = font
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER

            # 等级列颜色
            grade_cell = ws.cell(row=current_row, column=grade_column)
            grade_fill = _get_grade_fill(m.grade)
            if grade_fill:
                grade_cell.fill = grade_fill
            grade_cell.alignment = Alignment(horizontal="center")
            grade_cell.border = THIN_BORDER

            # 其他列边框
            for col in [1, 2, 9, 10]:
                cell = ws.cell(row=current_row, column=col)
                cell.border = THIN_BORDER
                if col in (1, 2):
                    cell.alignment = Alignment(horizontal="center")

        # --- 汇总行 ---
        if matches:
            summary_row = len(matches) + 2  # 表头+数据行之后
            ws.cell(row=summary_row, column=1, value="汇总")

            # 平均分
            avg_overall = sum(float(m.overall_score) for m in matches) / len(matches)
            ws.cell(row=summary_row, column=3, value=round(avg_overall, 2))

            # 各维度平均分
            for col, attr in zip(
                [4, 5, 6, 7],
                ["skill_score", "experience_score", "education_score", "soft_skill_score"],
            ):
                scores = [float(getattr(m, attr)) for m in matches if getattr(m, attr) is not None]
                if scores:
                    ws.cell(row=summary_row, column=col, value=round(sum(scores) / len(scores), 2))

            # 各等级人数
            grade_counts = {"excellent": 0, "qualified": 0, "unqualified": 0}
            for m in matches:
                if m.grade in grade_counts:
                    grade_counts[m.grade] += 1

            summary_text = (
                f"优秀: {grade_counts['excellent']}人 | "
                f"合格: {grade_counts['qualified']}人 | "
                f"不合格: {grade_counts['unqualified']}人 | "
                f"总计: {len(matches)}人"
            )
            ws.cell(row=summary_row, column=8, value=summary_text)

            # 汇总行样式
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=summary_row, column=col)
                cell.fill = SUMMARY_FILL
                cell.font = SUMMARY_FONT
                cell.border = THIN_BORDER

        # --- 自动列宽 ---
        _auto_adjust_column_width(ws)

        # --- 输出 ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_single_report_html(
        self,
        match_result: dict[str, Any],
        resume_data: dict[str, Any],
        job_data: dict[str, Any],
    ) -> str:
        """生成单份简历匹配报告（HTML格式）。

        生成结构化的HTML内容，前端可用 window.print() 打印为PDF。

        Args:
            match_result: 匹配结果字典，包含 overall_score, skill_score,
                         experience_score, education_score, soft_skill_score,
                         grade, recommendation, details
            resume_data: 简历解析数据（parsed_data JSONB内容）
            job_data: 岗位数据字典，包含 title, department, level, requirements

        Returns:
            完整的HTML字符串
        """
        # 安全转义
        job_title = escape(job_data.get("title", "未知岗位"))
        department = escape(job_data.get("department", ""))
        level = escape(job_data.get("level", ""))

        candidate_name = escape(resume_data.get("name", "未知候选人"))
        candidate_email = escape(resume_data.get("email", ""))
        candidate_phone = escape(resume_data.get("phone", ""))
        education = resume_data.get("education", [])
        experience = resume_data.get("experience", [])

        overall_score = float(match_result.get("overall_score", 0))
        skill_score = match_result.get("skill_score")
        experience_score = match_result.get("experience_score")
        education_score = match_result.get("education_score")
        soft_skill_score = match_result.get("soft_skill_score")
        grade = match_result.get("grade", "")
        recommendation = escape(match_result.get("recommendation", ""))
        details = match_result.get("details") or {}

        grade_label = GRADE_LABELS.get(grade, grade)
        grade_color = {
            "excellent": "#27ae60",
            "qualified": "#f39c12",
            "unqualified": "#e74c3c",
        }.get(grade, "#333")

        # 从 details 提取优劣势和详细分析
        strengths = details.get("strengths", [])
        weaknesses = details.get("weaknesses", [])
        dimension_analysis = details.get("dimension_analysis", {})

        # 雷达图数据（前端用）
        radar_data = {
            "labels": ["技能匹配", "经验匹配", "学历匹配", "软技能"],
            "values": [
                float(skill_score) if skill_score else 0,
                float(experience_score) if experience_score else 0,
                float(education_score) if education_score else 0,
                float(soft_skill_score) if soft_skill_score else 0,
            ],
        }

        # 生成教育经历HTML
        education_html = ""
        for edu in education[:5]:  # 最多5条
            school = escape(str(edu.get("school", "")))
            major = escape(str(edu.get("major", "")))
            degree = escape(str(edu.get("degree", "")))
            education_html += f"<li>{school} - {major} ({degree})</li>"

        # 生成工作经历HTML
        experience_html = ""
        for exp in experience[:5]:
            company = escape(str(exp.get("company", "")))
            position = escape(str(exp.get("position", "")))
            duration = escape(str(exp.get("duration", "")))
            experience_html += f"<li>{company} - {position} ({duration})</li>"

        # 优劣势列表
        strengths_html = "".join(f"<li>{escape(str(s))}</li>" for s in strengths)
        weaknesses_html = "".join(f"<li>{escape(str(w))}</li>" for w in weaknesses)

        # 维度详细分析
        dimension_html = ""
        for dim_key, dim_info in dimension_analysis.items():
            dim_name = escape(str(dim_key))
            dim_desc = escape(str(dim_info)) if isinstance(dim_info, str) else escape(
                str(dim_info.get("analysis", "")) if isinstance(dim_info, dict) else str(dim_info)
            )
            dimension_html += f"""
            <div class="dimension-item">
                <h4>{dim_name}</h4>
                <p>{dim_desc}</p>
            </div>"""

        generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # 预格式化分数显示值（避免f-string中条件格式说明符问题）
        skill_display = f"{float(skill_score):.1f}" if skill_score else "N/A"
        exp_display = f"{float(experience_score):.1f}" if experience_score else "N/A"
        edu_display = f"{float(education_score):.1f}" if education_score else "N/A"
        soft_display = f"{float(soft_skill_score):.1f}" if soft_skill_score else "N/A"

        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>匹配报告 - {job_title} - {candidate_name}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; line-height: 1.6; color: #333; padding: 40px; max-width: 900px; margin: 0 auto; }}
        .header {{ text-align: center; border-bottom: 2px solid #4472C4; padding-bottom: 20px; margin-bottom: 30px; }}
        .header h1 {{ color: #4472C4; font-size: 24px; }}
        .header .subtitle {{ color: #666; margin-top: 5px; }}
        .section {{ margin-bottom: 30px; }}
        .section h2 {{ color: #4472C4; font-size: 18px; border-left: 4px solid #4472C4; padding-left: 10px; margin-bottom: 15px; }}
        .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
        .info-item {{ padding: 8px 12px; background: #f8f9fa; border-radius: 4px; }}
        .info-item label {{ font-weight: bold; color: #555; }}
        .score-card {{ text-align: center; background: #f8f9fa; padding: 30px; border-radius: 8px; margin-bottom: 20px; }}
        .score-main {{ font-size: 48px; font-weight: bold; color: {grade_color}; }}
        .score-grade {{ font-size: 20px; color: {grade_color}; margin-top: 5px; }}
        .score-dimensions {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-top: 20px; }}
        .score-dim {{ text-align: center; padding: 15px; background: white; border-radius: 6px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
        .score-dim .value {{ font-size: 24px; font-weight: bold; }}
        .score-dim .label {{ font-size: 12px; color: #666; margin-top: 5px; }}
        .pros-cons {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .pros {{ background: #f0fff4; padding: 15px; border-radius: 6px; border-left: 4px solid #27ae60; }}
        .cons {{ background: #fff5f5; padding: 15px; border-radius: 6px; border-left: 4px solid #e74c3c; }}
        .pros h3 {{ color: #27ae60; }}
        .cons h3 {{ color: #e74c3c; }}
        ul {{ padding-left: 20px; margin-top: 10px; }}
        li {{ margin-bottom: 5px; }}
        .dimension-item {{ padding: 12px; background: #f8f9fa; border-radius: 4px; margin-bottom: 10px; }}
        .dimension-item h4 {{ color: #4472C4; margin-bottom: 5px; }}
        .recommendation {{ padding: 20px; background: #eef5fd; border-radius: 6px; font-size: 16px; }}
        .footer {{ text-align: center; color: #999; font-size: 12px; margin-top: 40px; padding-top: 20px; border-top: 1px solid #eee; }}
        .radar-data {{ display: none; }}
        @media print {{
            body {{ padding: 20px; }}
            .no-print {{ display: none; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>AI简历匹配报告</h1>
        <p class="subtitle">生成时间: {generated_at}</p>
    </div>

    <div class="section">
        <h2>岗位信息</h2>
        <div class="info-grid">
            <div class="info-item"><label>岗位名称：</label>{job_title}</div>
            <div class="info-item"><label>所属部门：</label>{department or "未指定"}</div>
            <div class="info-item"><label>职级：</label>{level or "未指定"}</div>
        </div>
    </div>

    <div class="section">
        <h2>候选人信息</h2>
        <div class="info-grid">
            <div class="info-item"><label>姓名：</label>{candidate_name}</div>
            <div class="info-item"><label>邮箱：</label>{candidate_email or "未提供"}</div>
            <div class="info-item"><label>电话：</label>{candidate_phone or "未提供"}</div>
        </div>
        {"<h3 style='margin-top:15px;'>教育经历</h3><ul>" + education_html + "</ul>" if education_html else ""}
        {"<h3 style='margin-top:15px;'>工作经历</h3><ul>" + experience_html + "</ul>" if experience_html else ""}
    </div>

    <div class="section">
        <h2>综合评分</h2>
        <div class="score-card">
            <div class="score-main">{overall_score:.1f}</div>
            <div class="score-grade">{grade_label}</div>
            <div class="score-dimensions">
                <div class="score-dim">
                    <div class="value">{skill_display}</div>
                    <div class="label">技能匹配</div>
                </div>
                <div class="score-dim">
                    <div class="value">{exp_display}</div>
                    <div class="label">经验匹配</div>
                </div>
                <div class="score-dim">
                    <div class="value">{edu_display}</div>
                    <div class="label">学历匹配</div>
                </div>
                <div class="score-dim">
                    <div class="value">{soft_display}</div>
                    <div class="label">软技能</div>
                </div>
            </div>
        </div>
        <!-- 雷达图数据供前端渲染 -->
        <div class="radar-data" id="radar-data" data-labels='{escape(",".join(radar_data["labels"]))}' data-values='{escape(",".join(str(v) for v in radar_data["values"]))}'></div>
    </div>

    {"<div class='section'><h2>优劣势分析</h2><div class='pros-cons'><div class='pros'><h3>✅ 优势</h3><ul>" + strengths_html + "</ul></div><div class='cons'><h3>⚠️ 不足</h3><ul>" + weaknesses_html + "</ul></div></div></div>" if strengths_html or weaknesses_html else ""}

    {"<div class='section'><h2>维度详细分析</h2>" + dimension_html + "</div>" if dimension_html else ""}

    <div class="section">
        <h2>推荐意见</h2>
        <div class="recommendation">{recommendation or "暂无推荐意见"}</div>
    </div>

    <div class="footer">
        <p>本报告由AI智能招聘系统自动生成，仅供参考。</p>
    </div>
</body>
</html>"""
        return html
